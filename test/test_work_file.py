import zipfile
from openpyxl import load_workbook
import csv
import pytest
from pypdf import PdfReader
from script_os import ZIP_DIR


def test_open_pdf(create_archive):
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('book1.pdf') as file_pdf:
            reader = PdfReader(file_pdf)
            text = reader.pages[0].extract_text()
            assert "alice@example.com" in text, f'Пользователь с таким email отсутсвует'

def test_open_csv(create_archive):
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('book1.csv') as file_csv:
            content = file_csv.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            first_row = csvreader[0]
            second_row = csvreader[2]
            assert 'АКТИВЕН' in first_row, f'Столбец статус "Активен" отсутсвует'
            assert 'FALSE' in second_row, f'Статус у пользователя не FALSE'

def test_open_xlsx(create_archive):
    with zipfile.ZipFile(ZIP_DIR) as zip_file:
        with zip_file.open('book1.xlsx') as file_xlsx:
            workbook = load_workbook(file_xlsx)
            sheet = workbook.active
            count_row = sheet.max_row
            assert count_row == 15, f'Количество строк в фале не соответсвует {count_row}'



